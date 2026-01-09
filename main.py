import os
import sys
import traci
import pandas as pd
import xml.etree.ElementTree as ET
from datetime import datetime
import json
from openpyxl.styles import Font, Alignment, PatternFill
from neural_traffic_controller import NeuralTrafficController


class SUMOCompetitionFramework:
    """
    SUMO竞赛数据收集框架

    框架结构:
    1. 环境初始化 (parse_config, parse_routes, initialize_environment)
    2. 控制算法实现 (apply_control_algorithm - 参赛者自定义)
    3. 数据收集与统计 (collect_step_data, save_to_excel)
    """

    def __init__(self, sumo_cfg_path):
        self.sumo_cfg_path = sumo_cfg_path
        self.routes_file = None
        self.net_file = None

        # 数据存储
        self.vehicle_data = []  # 车辆级数据
        self.step_data = []  # 时间步级数据
        self.route_data = {}  # 车辆路径数据
        self.vehicle_od_data = {}  # 车辆OD信息存储

        # 累计统计
        self.cumulative_departed = 0  # 累计出发车辆数
        self.cumulative_arrived = 0  # 累计到达车辆数
        self.all_departed_vehicles = set()  # 所有出发过的车辆
        self.all_arrived_vehicles = set()  # 所有到达过的车辆

        # 红绿灯监控
        self.traffic_lights = ['J5', 'J14', 'J15', 'J17']  # 可修改
        self.available_traffic_lights = []  # 实际可用的红绿灯

        # 仿真参数
        self.flow_rate = 0
        self.simulation_time = 0
        self.step_length = 1.0
        self.total_demand = 0  # 理论总需求

        print("=" * 70)
        print("SUMO竞赛数据收集框架")
        print("=" * 70)
        print("框架结构:")
        print("  第一部分: 环境初始化 (Baseline环境)")
        print("  第二部分: 控制算法 (参赛者自定义)")
        print("  第三部分: 数据统计与保存")
        print("=" * 70)

        # 初始化神经控制器
        self.neural_controller = None
        self.use_neural_control = True  # 启用神经控制

        # 加载控制器配置
        controller_config = {
            'node_dim': 9,
            'edge_dim': 4,
            'gnn_hidden_dim': 64,
            'gnn_output_dim': 256,
            'gnn_layers': 3,
            'gnn_heads': 4,
            'world_hidden_dim': 128,
            'future_steps': 5,
            'controller_hidden_dim': 128,
            'global_dim': 16,
            'top_k': 5,  # 只控制5辆车
            'ttc_threshold': 2.0,
            'thw_threshold': 1.5,
            'max_accel': 2.0,
            'max_decel': -3.0,
            'emergency_decel': -5.0,
            'max_lane_change_speed': 5.0,
            'cost_limit': 0.1,
            'lambda_lr': 0.01,
            'cache_timeout': 10,
            'device': 'cpu',  # 使用CPU以确保兼容性
            'model_path': None  # 预训练模型路径
        }

        try:
            self.neural_controller = NeuralTrafficController()
            print("✅ 神经交通控制器已加载")
        except Exception as e:
            print(f"❌ 神经控制器加载失败: {e}")
            self.use_neural_control = False

    # ========================================================================
    # 第一部分: 环境初始化 (Baseline环境)
    # ========================================================================

    def parse_config(self):
        """解析SUMO配置文件"""
        print("\n[第一部分] 正在初始化Baseline环境...")

        tree = ET.parse(self.sumo_cfg_path)
        root = tree.getroot()

        config_dir = os.path.dirname(self.sumo_cfg_path)

        # 获取路网和路径文件
        for input_elem in root.findall('.//input'):
            net_file = input_elem.find('net-file')
            if net_file is not None:
                net_file_path = net_file.get('value')
                if not os.path.isabs(net_file_path):
                    net_file_path = os.path.join(config_dir, net_file_path)
                self.net_file = net_file_path

            route_files = input_elem.find('route-files')
            if route_files is not None:
                route_file_path = route_files.get('value')
                if not os.path.isabs(route_file_path):
                    route_file_path = os.path.join(config_dir, route_file_path)
                self.routes_file = route_file_path

        # 获取时间步长
        time_step = root.find('.//step-length')
        if time_step is not None:
            self.step_length = float(time_step.get('value', 1.0))

        print(f"✓ 配置解析完成:")
        print(f"  - 网络文件: {self.net_file}")
        print(f"  - 路径文件: {self.routes_file}")
        print(f"  - 时间步长: {self.step_length}s")

    def parse_routes(self):
        """解析路径文件，计算车流量和总需求"""
        if not self.routes_file or not os.path.exists(self.routes_file):
            print("⚠️  路径文件不存在，无法计算理论需求")
            return

        try:
            tree = ET.parse(self.routes_file)
            root = tree.getroot()

            total_vehs_per_hour = 0
            max_end_time = 0
            total_demand = 0

            # 计算总需求
            for flow in root.findall('flow'):
                vehs_per_hour = float(flow.get('vehsPerHour', 0))
                begin_time = float(flow.get('begin', 0))
                end_time = float(flow.get('end', 0))

                duration_hours = (end_time - begin_time) / 3600.0
                flow_demand = vehs_per_hour * duration_hours
                total_demand += flow_demand

                total_vehs_per_hour += vehs_per_hour
                max_end_time = max(max_end_time, end_time)

            # 计算单独的trip数量
            trip_count = len(root.findall('trip'))
            total_demand += trip_count

            self.simulation_time = max_end_time
            self.flow_rate = total_vehs_per_hour / 3600.0
            self.total_demand = total_demand

            print(f"✓ 交通需求分析:")
            print(f"  - 流量率: {self.flow_rate:.4f} veh/s")
            print(f"  - 仿真时长: {self.simulation_time:.2f} s")
            print(f"  - 理论总需求: {self.total_demand:.0f} 车辆")
            print(f"  - 单独trip数量: {trip_count}")

        except Exception as e:
            print(f"❌ 路径文件解析失败: {e}")

    def initialize_traffic_lights(self):
        """初始化红绿灯监控"""
        try:
            all_tls = traci.trafficlight.getIDList()

            for tl_id in self.traffic_lights:
                if tl_id in all_tls:
                    self.available_traffic_lights.append(tl_id)
                else:
                    print(f"⚠️  红绿灯 {tl_id} 不存在于当前网络中")

            print(f"✓ 红绿灯监控设置:")
            print(f"  - 目标红绿灯: {self.traffic_lights}")
            print(f"  - 可用红绿灯: {self.available_traffic_lights}")
            print(f"  - 全部红绿灯: {list(all_tls)}")

        except Exception as e:
            print(f"❌ 红绿灯初始化失败: {e}")
            self.available_traffic_lights = []

    def initialize_environment(self, use_gui=True, max_steps=1000):
        """初始化SUMO仿真环境"""
        print("\n[第一部分] 正在启动SUMO仿真...")

        # 解析配置
        self.parse_config()
        self.parse_routes()

        # 启动SUMO
        sumo_binary = "sumo-gui" if use_gui else "sumo"
        sumo_cmd = [
            sumo_binary,
            "-c", self.sumo_cfg_path,
            "--no-warnings", "true",
            "--duration-log.statistics", "true"
        ]

        try:
            traci.start(sumo_cmd)
            print(f"✓ SUMO启动成功 (模式: {'GUI' if use_gui else 'CLI'})")
        except Exception as e:
            print(f"❌ SUMO启动失败: {e}")
            return False

        # 初始化红绿灯
        self.initialize_traffic_lights()

        print("✓ Baseline环境初始化完成!\n")
        return True

    # ========================================================================
    # 第二部分: 控制算法实现 (参赛者自定义)
    # ========================================================================

    def apply_control_algorithm(self, step):
        """
        应用控制优化算法 - 参赛者在此实现自己的算法

        参数:
            step: 当前仿真步数

        示例算法:
            - 自适应信号灯控制
            - 动态路径规划
            - 车辆速度控制
            - 交通流优化

        可用的TraCI函数示例:
            - traci.trafficlight.setPhase(tl_id, phase_index)
            - traci.trafficlight.setPhaseDuration(tl_id, duration)
            - traci.vehicle.setSpeed(veh_id, speed)
            - traci.vehicle.setRoute(veh_id, edge_list)
        """

        # ============================================================
        # 参赛者代码区域开始
        # ============================================================

        # 仅在特定步数执行控制以降低计算开销
        if step % 5 != 0:
            return

        if self.use_neural_control and self.neural_controller:
            try:
                # 收集当前车辆数据
                vehicle_data = self._collect_current_vehicle_data()

                if not vehicle_data:
                    return

                # 应用神经控制
                control_results = self.neural_controller.apply_control(vehicle_data, step)

                # 记录控制统计
                self._record_control_statistics(control_results, step)

            except Exception as e:
                print(f"⚠️  神经控制执行错误: {e}")
                # 回退到基础控制
                self._fallback_control_algorithm(step)
        else:
            # 回退到基础控制
            self._fallback_control_algorithm(step)

        # ============================================================
        # 参赛者代码区域结束
        # ============================================================

    def _collect_current_vehicle_data(self) -> Dict[str, Any]:
        """收集当前车辆数据"""
        vehicle_data = {}
        vehicle_ids = traci.vehicle.getIDList()

        for veh_id in vehicle_ids:
            try:
                # 确定是否为ICV (25%概率)
                is_icv = hash(veh_id) % 100 < 25

                # 获取车辆位置（简化）
                try:
                    position = traci.vehicle.getLanePosition(veh_id)
                except:
                    position = 0.0

                vehicle_data[veh_id] = {
                    'position': position,
                    'speed': traci.vehicle.getSpeed(veh_id),
                    'acceleration': traci.vehicle.getAcceleration(veh_id),
                    'lane_index': traci.vehicle.getLaneIndex(veh_id),
                    'remaining_distance': 1000.0,  # 简化
                    'completion_rate': 0.5,  # 简化
                    'is_icv': is_icv,
                    'id': veh_id,
                    'lane_id': traci.vehicle.getLaneID(veh_id)
                }
            except Exception as e:
                continue

        return vehicle_data

    def _fallback_control_algorithm(self, step):
        """回退控制算法"""
        # 简单的自适应速度控制
        vehicle_ids = traci.vehicle.getIDList()

        for veh_id in vehicle_ids:
            try:
                current_speed = traci.vehicle.getSpeed(veh_id)
                edge_id = traci.vehicle.getRoadID(veh_id)

                # 仅控制25%的车辆（模拟ICV）
                if hash(veh_id) % 4 != 0:  # 25%概率
                    continue

                # 基于前方车辆数的速度控制
                leader_count = 0
                try:
                    next_edge = traci.vehicle.getRoute(veh_id)[traci.vehicle.getRouteIndex(veh_id) + 1]
                    leader_count = traci.edge.getLastStepVehicleNumber(next_edge)
                except:
                    leader_count = 0

                # 调整速度
                if leader_count > 5:
                    target_speed = max(5.0, current_speed - 1.0)
                else:
                    target_speed = min(25.0, current_speed + 0.5)

                traci.vehicle.setSpeedMode(veh_id, 0)  # 关闭SUMO自动控制
                traci.vehicle.setSpeed(veh_id, target_speed)

            except Exception as e:
                continue

    def _record_control_statistics(self, control_results: Dict[str, Any], step: int):
        """记录控制统计"""
        if not hasattr(self, 'control_stats'):
            self.control_stats = {
                'total_interventions': 0,
                'total_emergency_interventions': 0,
                'total_controlled_vehicles': 0,
                'step_records': []
            }

        self.control_stats['total_interventions'] += control_results['safety_interventions']
        self.control_stats['total_emergency_interventions'] += control_results['emergency_interventions']
        self.control_stats['total_controlled_vehicles'] += len(control_results['controlled_vehicles'])

        # 每100步记录一次
        if step % 100 == 0:
            record = {
                'step': step,
                'controlled_vehicles': len(control_results['controlled_vehicles']),
                'interventions': control_results['safety_interventions'],
                'emergency_interventions': control_results['emergency_interventions']
            }
            self.control_stats['step_records'].append(record)

    # ========================================================================
    # 第三部分: 数据收集与统计
    # ========================================================================

    def get_traffic_light_states(self):
        """获取红绿灯状态"""
        tl_states = {}

        for tl_id in self.available_traffic_lights:
            try:
                state = traci.trafficlight.getRedYellowGreenState(tl_id)
                phase = traci.trafficlight.getPhase(tl_id)
                remaining_time = traci.trafficlight.getNextSwitch(tl_id) - traci.simulation.getTime()

                tl_states[f'{tl_id}_state'] = state
                tl_states[f'{tl_id}_phase'] = phase
                tl_states[f'{tl_id}_remaining_time'] = remaining_time

            except Exception as e:
                tl_states[f'{tl_id}_state'] = 'unknown'
                tl_states[f'{tl_id}_phase'] = -1
                tl_states[f'{tl_id}_remaining_time'] = -1

        return tl_states

    def get_vehicle_od(self, veh_id):
        """获取车辆OD信息"""
        if veh_id in self.vehicle_od_data:
            return self.vehicle_od_data[veh_id]

        try:
            route = traci.vehicle.getRoute(veh_id)
            if len(route) >= 2:
                origin = route[0]
                destination = route[-1]
            elif len(route) == 1:
                origin = route[0]
                destination = route[0]
            else:
                origin = "unknown"
                destination = "unknown"

            od_info = {
                'origin': origin,
                'destination': destination,
                'route_length': len(route)
            }

            self.vehicle_od_data[veh_id] = od_info
            return od_info

        except:
            od_info = {
                'origin': "unknown",
                'destination': "unknown",
                'route_length': 0
            }
            self.vehicle_od_data[veh_id] = od_info
            return od_info

    def get_route_length(self, edges):
        """计算路径总长度"""
        total_length = 0
        for edge_id in edges:
            try:
                edge_length = traci.edge.getLength(edge_id)
                total_length += edge_length
            except:
                try:
                    lane_id = f"{edge_id}_0"
                    edge_length = traci.lane.getLength(lane_id)
                    total_length += edge_length
                except:
                    total_length += 100
        return total_length

    def calculate_traveled_distance(self, veh_id, route_info):
        """计算车辆已行驶距离"""
        try:
            current_edge = traci.vehicle.getRoadID(veh_id)
            current_position = traci.vehicle.getLanePosition(veh_id)
            route_edges = route_info['route_edges']

            traveled = 0
            for edge in route_edges:
                if edge == current_edge:
                    traveled += current_position
                    break
                else:
                    try:
                        edge_length = traci.edge.getLength(edge)
                        traveled += edge_length
                    except:
                        traveled += 100

            return min(traveled, route_info['route_length'])
        except:
            return 0

    def collect_step_data(self, step):
        """收集每个时间步的数据"""
        current_time = step * self.step_length

        # 获取当前活跃车辆
        current_vehicle_ids = set(traci.vehicle.getIDList())

        # 更新累计统计
        current_arrived_ids = set(traci.simulation.getArrivedIDList())
        current_departed_ids = set(traci.simulation.getDepartedIDList())

        new_arrivals = current_arrived_ids - self.all_arrived_vehicles
        self.all_arrived_vehicles.update(new_arrivals)
        self.cumulative_arrived = len(self.all_arrived_vehicles)

        new_departures = current_departed_ids - self.all_departed_vehicles
        self.all_departed_vehicles.update(new_departures)
        self.cumulative_departed = len(self.all_departed_vehicles)

        # 获取红绿灯状态
        traffic_light_states = self.get_traffic_light_states()

        # 记录时间步级数据
        step_record = {
            'step': step,
            'time': current_time,
            'active_vehicles': len(current_vehicle_ids),
            'arrived_vehicles': self.cumulative_arrived,
            'departed_vehicles': self.cumulative_departed,
            'current_arrivals': len(new_arrivals),
            'current_departures': len(new_departures)
        }
        step_record.update(traffic_light_states)
        self.step_data.append(step_record)

        # 收集车辆级数据
        for veh_id in current_vehicle_ids:
            try:
                speed = traci.vehicle.getSpeed(veh_id)
                position = traci.vehicle.getLanePosition(veh_id)
                edge_id = traci.vehicle.getRoadID(veh_id)
                route_index = traci.vehicle.getRouteIndex(veh_id)

                od_info = self.get_vehicle_od(veh_id)

                if veh_id not in self.route_data:
                    route_edges = traci.vehicle.getRoute(veh_id)
                    route_length = self.get_route_length(route_edges)
                    self.route_data[veh_id] = {
                        'route_edges': route_edges,
                        'route_length': route_length
                    }

                route_info = self.route_data[veh_id]
                traveled_distance = self.calculate_traveled_distance(veh_id, route_info)
                completion_rate = min(traveled_distance / max(route_info['route_length'], 1), 1.0)

                vehicle_record = {
                    'step': step,
                    'time': current_time,
                    'vehicle_id': veh_id,
                    'speed': speed,
                    'position': position,
                    'edge_id': edge_id,
                    'route_index': route_index,
                    'traveled_distance': traveled_distance,
                    'route_length': route_info['route_length'],
                    'completion_rate': completion_rate,
                    'origin': od_info['origin'],
                    'destination': od_info['destination'],
                    'route_edges_count': od_info['route_length']
                }
                self.vehicle_data.append(vehicle_record)

            except Exception as e:
                continue

        # 进度报告
        if step % 100 == 0:
            print(f"[步骤 {step}] 活跃: {len(current_vehicle_ids)}, "
                  f"累计出发: {self.cumulative_departed}, "
                  f"累计到达: {self.cumulative_arrived}")

    def save_to_excel(self, output_dir="competition_results"):
        """保存数据到Excel文件 - 替代原来的save_to_csv方法"""
        print(f"\n[第三部分] 正在保存优化后的仿真数据到Excel...")

        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # 生成Excel文件名
        excel_file = os.path.join(output_dir, f"submit.xlsx")

        # 准备仿真参数
        params = {
            'flow_rate': self.flow_rate,
            'simulation_time': self.simulation_time,
            'step_length': self.step_length,
            'total_steps': len(self.step_data),
            'total_demand': self.total_demand,
            'final_departed': self.cumulative_departed,
            'final_arrived': self.cumulative_arrived,
            'unique_vehicles': len(self.route_data),
            'monitored_traffic_lights': self.traffic_lights,
            'available_traffic_lights': self.available_traffic_lights,
            'collection_timestamp': timestamp
        }

        # 如果有控制统计数据，也加入参数中
        if hasattr(self, 'control_stats'):
            params.update({
                'total_controlled_vehicles': self.control_stats['total_controlled_vehicles'],
                'total_interventions': self.control_stats['total_interventions'],
                'total_emergency_interventions': self.control_stats['total_emergency_interventions'],
            })

        # 创建Excel写入器
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:

            # 1. 写入数据汇总 Sheet
            print(f"✓ 正在创建数据汇总...")
            summary_data = [
                {'指标': '理论总需求', '数值': f"{params['total_demand']:.0f} 车辆"},
                {'指标': '实际累计出发', '数值': f"{params['final_departed']} 车辆"},
                {'指标': '实际累计到达', '数值': f"{params['final_arrived']} 车辆"},
                {'指标': '唯一车辆数', '数值': f"{params['unique_vehicles']} 车辆"},
                {'指标': '总时间步数', '数值': params['total_steps']},
                {'指标': '仿真时长', '数值': f"{params['simulation_time']:.2f} 秒"},
                {'指标': '流量率', '数值': f"{params['flow_rate']:.4f} veh/s"},
                {'指标': '时间步长', '数值': f"{params['step_length']:.2f} 秒"},
                {'指标': '监控红绿灯', '数值': ', '.join(params['available_traffic_lights'])},
                {'指标': '数据收集时间', '数值': timestamp}
            ]

            # 如果有控制统计数据，也添加到汇总中
            if hasattr(self, 'control_stats'):
                summary_data.extend([
                    {'指标': '受控车辆总数', '数值': f"{params['total_controlled_vehicles']} 车辆"},
                    {'指标': '安全干预总数', '数值': f"{params['total_interventions']} 次"},
                    {'指标': '紧急干预总数', '数值': f"{params['total_emergency_interventions']} 次"}
                ])

            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='数据汇总', index=False)

            # 格式化汇总sheet
            worksheet = writer.sheets['数据汇总']
            worksheet.column_dimensions['A'].width = 25
            worksheet.column_dimensions['B'].width = 40

            for cell in worksheet[1]:
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 2. 写入仿真参数 Sheet
            print(f"✓ 正在写入仿真参数...")
            params_data = []
            for key, value in params.items():
                if isinstance(value, list):
                    value = ', '.join(map(str, value))
                params_data.append({'参数名称': key, '参数值': str(value)})

            params_df = pd.DataFrame(params_data)
            params_df.to_excel(writer, sheet_name='仿真参数', index=False)

            # 格式化参数sheet
            worksheet = writer.sheets['仿真参数']
            worksheet.column_dimensions['A'].width = 30
            worksheet.column_dimensions['B'].width = 50

            for cell in worksheet[1]:
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 3. 写入时间步数据 Sheet
            if self.step_data:
                print(f"✓ 正在写入时间步数据...")
                step_df = pd.DataFrame(self.step_data)
                step_df.to_excel(writer, sheet_name='时间步数据', index=False)

                # 格式化
                worksheet = writer.sheets['时间步数据']
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, size=11, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                print(f"  - 记录数: {len(step_df):,}")

                tl_columns = [col for col in step_df.columns if
                              any(tl_id in col for tl_id in self.available_traffic_lights)]
                if tl_columns:
                    print(f"  - 红绿灯数据列: {len(tl_columns)}")

            # 4. 写入车辆数据 Sheet
            if self.vehicle_data:
                print(f"✓ 正在写入车辆数据...")
                vehicle_df = pd.DataFrame(self.vehicle_data)
                vehicle_df.to_excel(writer, sheet_name='车辆数据', index=False)

                # 格式化
                worksheet = writer.sheets['车辆数据']
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, size=11, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                print(f"  - 记录数: {len(vehicle_df):,}")
                print(f"  - 唯一车辆: {vehicle_df['vehicle_id'].nunique()}")

                if 'origin' in vehicle_df.columns:
                    unique_od_pairs = vehicle_df.groupby('vehicle_id')[['origin', 'destination']].first()
                    print(f"  - 唯一OD对: {len(unique_od_pairs.drop_duplicates())}")

        # 数据统计报告
        print(f"\n{'=' * 70}")
        print(f"数据收集统计报告")
        print(f"{'=' * 70}")
        print(f"理论总需求:     {self.total_demand:.0f} 车辆")
        print(f"实际累计出发:   {self.cumulative_departed} 车辆")
        print(f"实际累计到达:   {self.cumulative_arrived} 车辆")
        
        if hasattr(self, 'control_stats'):
            print(f"受控车辆总数:   {self.control_stats['total_controlled_vehicles']} 车辆")
            print(f"安全干预总数:   {self.control_stats['total_interventions']} 次")
            print(f"紧急干预总数:   {self.control_stats['total_emergency_interventions']} 次")
        print(f"{'=' * 70}")
        
        print(f"\n✅ Excel文件已保存: {excel_file}")
        print(f"\nExcel包含以下sheets:")
        print(f"  1. 数据汇总 - 关键指标统计")
        print(f"  2. 仿真参数 - 完整的仿真配置参数")
        print(f"  3. 时间步数据 - 每个时间步的系统状态")
        print(f"  4. 车辆数据 - 每辆车每个时间步的详细数据")

        return {'excel_file': excel_file}

    def run(self, max_steps=3600, use_gui=True):
        """运行完整的仿真流程"""
        print("\n开始运行SUMO竞赛仿真框架...")
        print(f"最大步数: {max_steps}\n")

        # 第一部分: 初始化Baseline环境
        if not self.initialize_environment(use_gui=use_gui, max_steps=max_steps):
            print("❌ 环境初始化失败")
            return False

        # 仿真主循环
        print(f"\n{'=' * 70}")
        print("[第二部分] 开始运行控制算法...")
        print(f"{'=' * 70}\n")

        step = 0
        try:
            while step < max_steps:
                # 执行仿真步
                traci.simulationStep()

                # 第二部分: 应用控制算法
                self.apply_control_algorithm(step)

                # 第三部分: 收集数据
                self.collect_step_data(step)

                step += 1

                # 检查仿真是否结束
                if traci.simulation.getMinExpectedNumber() <= 0 and step > 100:
                    print(f"\n仿真自然结束于步骤 {step}")
                    break

        except Exception as e:
            print(f"\n❌ 仿真过程中发生错误: {e}")
            import traceback
            traceback.print_exc()

        finally:
            traci.close()

        # 第三部分: 保存数据到Excel
        print(f"\n{'=' * 70}")
        result = self.save_to_excel()

        print(f"\n✅ 仿真完成!")
        print(f"\n可使用此Excel文件进行评测提交: {result['excel_file']}")

        return True


def main():
    """主函数 - 参赛者使用入口"""

    # ========================================================================
    # 配置区域 - 参赛者修改此处
    # ========================================================================

    # 方式1: 从命令行参数获取配置文件路径
    if len(sys.argv) > 1:
        sumo_cfg = sys.argv[1]
    else:
        # 方式2: 直接指定配置文件路径
        sumo_cfg = r"仿真环境-初赛\sumo.sumocfg"

    # 仿真参数设置
    MAX_STEPS = 3600 # 最大仿真步数
    USE_GUI = True  # 是否使用GUI界面

    # ========================================================================

    # 检查配置文件是否存在
    if not os.path.exists(sumo_cfg):
        print(f"❌ 配置文件不存在: {sumo_cfg}")
        print("\n请修改main()函数中的sumo_cfg路径，或使用命令行参数:")
        print(f"python {sys.argv[0]} <your_config_file.sumocfg>")
        return

    try:
        # 创建框架实例
        framework = SUMOCompetitionFramework(sumo_cfg)

        # 运行仿真
        framework.run(max_steps=MAX_STEPS, use_gui=USE_GUI)

    except Exception as e:
        print(f"\n❌ 程序运行失败: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()